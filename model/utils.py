from torch.optim import Adam
from torch.optim.lr_scheduler import CosineAnnealingLR
from sklearn.metrics import confusion_matrix, precision_score, recall_score, f1_score
import pandas as pd
import os
import torch
import numpy as np
import matplotlib.pyplot as plt
import rasterio
from openpyxl.drawing.image import Image as ExcelImage
from matplotlib import pyplot as plt
import rasterio
from openpyxl import load_workbook
from PIL import Image as PILImage

def build_optimizer(config, model_params):
    """Build optimizer based on configuration."""
    optimizer_type = config.training["optimizer"]["type"]
    optimizer_params = config.training["optimizer"]["params"]
    if optimizer_type == "Adam":
        return Adam(model_params, **optimizer_params)
    else:
        raise ValueError(f"Unsupported optimizer: {optimizer_type}")

def build_scheduler(config, optimizer, train_loader_length, current_epoch):
    """Build scheduler based on configuration."""
    scheduler_type = config.training["scheduler"]["type"]
    T_max_factor = config.training["scheduler"]["T_max_factor"]
    if scheduler_type == "CosineAnnealingLR":
        T_max = int(T_max_factor * (config.training["epochs"] - current_epoch) * train_loader_length)
        return CosineAnnealingLR(optimizer, T_max=T_max)
    else:
        raise ValueError(f"Unsupported scheduler: {scheduler_type}")

def _compute_metrics(all_labels, all_preds):
    """Compute precision, recall, f1 and confusion matrix"""
    tn, fp, fn, tp = confusion_matrix(all_labels, all_preds, labels=[0, 1]).ravel()
    precision = precision_score(all_labels, all_preds, average=None, zero_division=0, labels=[0, 1])
    recall = recall_score(all_labels, all_preds, average=None, zero_division=0, labels=[0, 1])
    f1 = f1_score(all_labels, all_preds, average=None, zero_division=0, labels=[0, 1])
    return {
        'TN': tn, 'FP': fp, 'FN': fn, 'TP': tp,
        'Precision_0': precision[0], 'Recall_0': recall[0], 'F1_0': f1[0],
        'Precision_1': precision[1], 'Recall_1': recall[1], 'F1_1': f1[1],
    }

def compute_epoch_metrics(labels, preds, loss, elapsed_time):
    """Compute metrics for the entire epoch"""
    core_metrics = _compute_metrics(labels, preds)
    metrics = {
        **core_metrics,
        'Loss': loss,
        'Time': elapsed_time
    }
    return metrics

def compute_image_metrics(labels, preds, image_id):
    """Compute metrics for a single image"""
    core_metrics = _compute_metrics(labels, preds)
    metrics = {
        'Image_ID': image_id,
        **core_metrics
    }
    return metrics

def log_epoch_results(epoch, train_metrics, val_metrics, test_metrics, experiment_name, results_path):
    """Log metrics for current epoch into an Excel sheet"""
    
    # Create row with metrics
    row = {
        'Epoch': epoch,
        # Train
        'Train_Loss': f"{train_metrics['Loss']:.4f}",
        'Train_Precision_1': f"{train_metrics['Precision_1']:.4f}", 'Train_Recall_1': f"{train_metrics['Recall_1']:.4f}", 'Train_F1_1': f"{train_metrics['F1_1']:.4f}",
        'Train_Precision_0': f"{train_metrics['Precision_0']:.4f}", 'Train_Recall_0': f"{train_metrics['Recall_0']:.4f}", 'Train_F1_0': f"{train_metrics['F1_0']:.4f}", 'Train_TN': f"{train_metrics['TN']:.4f}", 'Train_FP': f"{train_metrics['FP']:.4f}", 'Train_FN': f"{train_metrics['FN']:.4f}", 'Train_TP': f"{train_metrics['TP']:.4f}",
        'Train_Time': train_metrics['Time'].split(".")[0],
        # Val
        'Val_Loss': f"{val_metrics['Loss']:.4f}",
        'Val_Precision_1': f"{val_metrics['Precision_1']:.4f}", 'Val_Recall_1': f"{val_metrics['Recall_1']:.4f}", 'Val_F1_1': f"{val_metrics['F1_1']:.4f}",
        'Val_Precision_0': f"{val_metrics['Precision_0']:.4f}", 'Val_Recall_0': f"{val_metrics['Recall_0']:.4f}", 'Val_F1_0': f"{val_metrics['F1_0']:.4f}", 'Val_TN': f"{val_metrics['TN']:.4f}", 'Val_FP': f"{val_metrics['FP']:.4f}", 'Val_FN': f"{val_metrics['FN']:.4f}", 'Val_TP': f"{val_metrics['TP']:.4f}",
        'Val_Time': val_metrics['Time'].split(".")[0],
        # Test
        'Test_Loss': f"{test_metrics['Loss']:.4f}",
        'Test_Precision_1': f"{test_metrics['Precision_1']:.4f}", 'Test_Recall_1': f"{test_metrics['Recall_1']:.4f}", 'Test_F1_1': f"{test_metrics['F1_1']:.4f}",
        'Test_Precision_0': f"{test_metrics['Precision_0']:.4f}", 'Test_Recall_0': f"{test_metrics['Recall_0']:.4f}", 'Test_F1_0': f"{test_metrics['F1_0']:.4f}", 'Test_TN': f"{test_metrics['TN']:.4f}", 'Test_FP': f"{test_metrics['FP']:.4f}", 'Test_FN': f"{test_metrics['FN']:.4f}", 'Test_TP': f"{test_metrics['TP']:.4f}",
        'Test_Time': test_metrics['Time'].split(".")[0],
    }

    # Create DataFrame and append to Excel file
    df_new = pd.DataFrame([row])
    if os.path.exists(results_path):
        try:
            # Try to append to the existing sheet
            existing_df = pd.read_excel(results_path, sheet_name=experiment_name)
            full_df = pd.concat([existing_df, df_new], ignore_index=True)
        except ValueError:
            # The sheet does not exist, create it
            full_df = df_new
        # Write to the existing file
        with pd.ExcelWriter(results_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            full_df.to_excel(writer, sheet_name=experiment_name, index=False)
    else:
        # The file does not exist, create it
        with pd.ExcelWriter(results_path, engine='openpyxl') as writer:
            df_new.to_excel(writer, sheet_name=experiment_name, index=False)

def compute_sam(v1, v2):
    """Compute the Spectral Angle Mapper (SAM) between two sets of vectors."""
    dot = np.sum(v1 * v2, axis=1)
    norm1 = np.linalg.norm(v1, axis=1)
    norm2 = np.linalg.norm(v2, axis=1)
    cos_angle = np.clip(dot / (norm1 * norm2 + 1e-10), -1.0, 1.0)
    return np.arccos(cos_angle)

def reconstruct_image(config, img_id, positions, values):
    """Reconstruct the image from the predicted labels."""
    reference_path = os.path.join(
        config.paths["sentinel_data_dir_2020"],
        config.filenames["images"].format(id=img_id)
    )
    with rasterio.open(reference_path) as src:
        _, h, w = src.read().shape
    image = np.zeros((h, w), dtype=np.uint8)
    for (r, c), v in zip(positions, values):
        image[r, c] = v
    return image

def save_prediction_image(config, output_dir, filename_format, image_id, array):
    """Save the predicted image."""
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(
        output_dir,
        filename_format.format(id=image_id)
    )

    # Copy the profile from the reference image
    reference_path = os.path.join(
        config.paths["sentinel_data_dir_2020"],
        config.filenames["images"].format(id=image_id)
    )
    with rasterio.open(reference_path) as src:
        profile = src.profile

    # Update the profile for the output image and save it
    profile.update(
        {
            "count": 1,
            "dtype": array.dtype,
            "nodata": 0
        }
    )
    with rasterio.open(out_path, 'w', **profile) as dst:
        dst.write(array, 1)
        
def insert_image_column(ws, df, image_base_path, filename_format, column_letter, column_name, temp_pngs):
    ws[f"{column_letter}1"] = column_name
    for i, row in df.iterrows():
        image_id = row["Image_ID"]
        tif_path = os.path.join(image_base_path, filename_format.format(id=image_id))
        png_path = tif_path.replace(".tif", ".png")

        if os.path.exists(tif_path):
            with rasterio.open(tif_path) as src:
                array = src.read(1)
            plt.imsave(png_path, array, cmap="tab10")
            temp_pngs.add(png_path)

            if os.path.exists(png_path):
                with PILImage.open(png_path) as original:
                    orig_width, orig_height = original.size
                target_height = 150
                scale = target_height / orig_height
                target_width = int(orig_width * scale)

                img = ExcelImage(png_path)
                img.height = target_height
                img.width = target_width

                excel_row = i + 2
                ws.row_dimensions[excel_row].height = 125
                ws.add_image(img, f"{column_letter}{excel_row}")

def insert_images_into_excel(writer_path, df, sheet_name, results_base_path, result_filename_format, ground_truth_base_path, ground_truth_filename_format):
    """Insert result and ground truth images into the Excel file."""
    wb = load_workbook(writer_path)
    temp_pngs = set()

    ws = wb[sheet_name]
    insert_image_column(
        ws,
        df,
        results_base_path,
        result_filename_format,
        "L",
        "Predictions",
        temp_pngs
    )
    insert_image_column(
        ws,
        df,
        ground_truth_base_path,
        ground_truth_filename_format,
        "M",
        "Ground_truth",
        temp_pngs
    )

    wb.save(writer_path)

    for path in temp_pngs:
        os.remove(path)
